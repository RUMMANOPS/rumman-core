#!/usr/bin/env python3
"""
generate_ops_workbook.py — RUMMAN master operational workbook.

Produces a single xlsx file that is the complete operating system inventory
for RUMMAN. Designed to survive context loss: a competent new person opens
this file and understands the platform in < 30 minutes.

Covers: architecture, accounts, costs, corpus, strategy, risks, open decisions.

Output: data/RUMMAN_OPS_YYYY-MM-DD.xlsx  (or --out path)

Usage:
    python3 scripts/generate_ops_workbook.py
    python3 scripts/generate_ops_workbook.py --out ~/Desktop/RUMMAN_OPS.xlsx

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


def _load_env(repo_root: Path) -> dict:
    """Parse .env from repo root — values populate the KEYS sheet at generation time."""
    env_path = repo_root / ".env"
    result = {}
    if not env_path.exists():
        return result
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, _, v = line.partition("=")
            result[k.strip()] = v.strip()
    return result


# ── Colour palette ─────────────────────────────────────────────────────────────

C_NAVY    = "1A1A2E"   # title backgrounds
C_DARK    = "16213E"   # section headers
C_MID     = "0F3460"   # column headers
C_ACCENT  = "E94560"   # highlights / critical markers
C_WHITE   = "FFFFFF"
C_LGRAY   = "F5F5F5"   # alternating row tint
C_MGRAY   = "DCDCDC"   # borders / dividers
C_GREEN   = "27AE60"
C_YELLOW  = "F39C12"
C_RED     = "E74C3C"
C_BLUE    = "2980B9"
C_PURPLE  = "8E44AD"

TAB_COLORS = {
    "COMPASS":   "E94560",
    "SYSTEMS":   "2980B9",
    "PROCESSES": "27AE60",
    "CORPUS":    "8E44AD",
    "STRATEGY":  "F39C12",
    "OPEN ITEMS":"E74C3C",
    "KEYS":      "922B21",
}

C_CRED_DARK = "7B241C"   # dark red for credential subtitle bar


# ── Style factories ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=C_WHITE, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin", color=C_MGRAY) -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color=C_MGRAY) -> Border:
    s = Side(style="thin", color=color)
    return Border(bottom=s)


# ── Low-level cell writers ─────────────────────────────────────────────────────

def title_cell(ws, row, col, text, span=None,
               bg=C_NAVY, fg=C_WHITE, size=14, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=bold, size=size, color=fg)
    c.fill = _fill(bg)
    c.alignment = _align("left", "center")
    if span:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
    return c


def section_header(ws, row, col, text, span=None, bg=C_DARK):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, size=10, color=C_WHITE)
    c.fill = _fill(bg)
    c.alignment = _align("left", "center")
    if span:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
    return c


def col_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, size=9, color=C_WHITE)
    c.fill = _fill(C_MID)
    c.alignment = _align("center", "center", wrap=True)
    c.border = _border()
    return c


def data_cell(ws, row, col, value, bold=False, color=None,
              bg=None, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=9,
                  color=color or "000000")
    if bg:
        c.fill = _fill(bg)
    elif row % 2 == 0:
        c.fill = _fill(C_LGRAY)
    c.alignment = _align(align, "center", wrap=wrap)
    c.border = _border(style="hair")
    return c


def status_cell(ws, row, col, value):
    colors = {
        "LIVE":    (C_GREEN, C_WHITE),
        "GATED":   (C_YELLOW, "000000"),
        "PLANNED": (C_BLUE, C_WHITE),
        "DONE":    (C_GREEN, C_WHITE),
        "PENDING": (C_YELLOW, "000000"),
        "APPLIED": (C_GREEN, C_WHITE),
        "FAILED":  (C_RED, C_WHITE),
        "ACTIVE":  (C_GREEN, C_WHITE),
        "YES":     (C_RED, C_WHITE),
        "NO":      (C_GREEN, C_WHITE),
        "HIGH":    (C_RED, C_WHITE),
        "MEDIUM":  (C_YELLOW, "000000"),
        "LOW":     (C_GREEN, C_WHITE),
        "OPEN":    (C_YELLOW, "000000"),
        "CLOSED":  (C_GREEN, C_WHITE),
        "CURRENT": (C_ACCENT, C_WHITE),
    }
    bg, fg = colors.get(str(value).upper().strip(), (None, "000000"))
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=9, color=fg)
    if bg:
        c.fill = _fill(bg)
    elif row % 2 == 0:
        c.fill = _fill(C_LGRAY)
    c.alignment = _align("center", "center")
    c.border = _border(style="hair")
    return c


def blank_row(ws, row, span, bg=C_WHITE):
    for col in range(1, span + 1):
        c = ws.cell(row=row, column=col, value=None)
        c.fill = _fill(bg)


# ── Sheet 1: COMPASS ───────────────────────────────────────────────────────────

def build_compass(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["COMPASS"]
    ws.freeze_panes = "A2"

    COLS = 8
    r = 1

    # ── Title block
    title_cell(ws, r, 1, "  رمّان  ·  RUMMAN — Master Operations Inventory",
               span=COLS, bg=C_NAVY, size=16)
    r += 1
    title_cell(ws, r, 1,
               f"  Generated: {date.today().isoformat()}   ·   Repository: rumman-core @ github.com/RUMMANOPS/rumman-core",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 2

    # ── What RUMMAN Is
    section_header(ws, r, 1, "  WHAT IS RUMMAN", span=COLS)
    r += 1
    for line in [
        ("Operational intelligence platform for Saudi university students.", True),
        ("Ingests institutional knowledge (study plans, regulations, course descriptions) and live community", False),
        ("intelligence (148K+ Telegram messages from student groups) to deliver grounded academic assistance.", False),
        ("The corpus — not the AI — is the product. AI is the retrieval and synthesis lens.", True),
    ]:
        c = ws.cell(row=r, column=1, value=f"  {line[0]}")
        c.font = Font(bold=line[1], size=10, color="000000")
        c.fill = _fill(C_LGRAY)
        c.alignment = _align("left", "center", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
        r += 1
    r += 1

    # ── What RUMMAN Is NOT
    section_header(ws, r, 1, "  WHAT RUMMAN IS NOT", span=COLS)
    r += 1
    nots = [
        "NOT a chatbot that invents answers — every response is grounded in the ingested corpus.",
        "NOT a replacement for the university — a companion that makes existing resources accessible.",
        "NOT a search engine — it synthesizes intelligence across multiple source types.",
        "NOT an LLM wrapper — the differentiation is the community corpus, not the model.",
        "NOT a content creator — it surfaces what students and institutions have already produced.",
    ]
    for n in nots:
        c = ws.cell(row=r, column=1, value=f"  ✕  {n}")
        c.font = Font(size=9, color="000000")
        c.fill = _fill(C_LGRAY)
        c.alignment = _align("left", "center", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
        r += 1
    r += 1

    # ── System Status
    section_header(ws, r, 1, "  LIVE SYSTEM STATUS  (June 2026)", span=COLS)
    r += 1
    headers = ["Process", "File", "Role", "Status", "Account", "Gate Variable", "Notes"]
    for i, h in enumerate(headers):
        col_header(ws, r, i + 1, h)
    r += 1
    processes = [
        ("listener",      "rumman_engine.py",              "Live Telegram NewMessage → messages table",         "LIVE",   "غيث",     "",                               "Never calls iter_messages (backfill guard)"),
        ("backfill",      "telegram_backfill_worker.py",   "Claims backfill_jobs, ingests history w/ lease",    "LIVE",   "راوي",    "",                               "Gap-fill support; discovers new chats"),
        ("media",         "telegram_download_worker.py",   "Unified audio_transcribe + telegram_media jobs",    "LIVE",   "إبراهيم", "",                               "~11K failed jobs pending reset after session fix"),
        ("embed",         "embed_worker.py",               "embed_chunk jobs → OpenAI → document_chunks",       "LIVE",   "—",       "",                               "Question-aware chunking; NFKC normalization"),
        ("search",        "search_api.py",                 "FastAPI: query understanding → pgvector → synthesis","LIVE",  "—",       "",                               "LRU cache 2h TTL; course-specific queries only"),
        ("bot",           "telegram_bot.py",               "Student-facing bot; calls /synthesize",             "LIVE",   "Bot token","",                               "Enrollment persistence; 6-turn conversation history"),
        ("intelligence",  "intelligence_worker.py",        "Extract assignments/deadlines from messages",       "GATED",  "—",       "INTELLIGENCE_WORKER_ENABLED=true","Max 200K tokens/run; cursor-tracked"),
        ("attribution",   "attribution_worker.py",         "AI course attribution for untagged chunks",         "GATED",  "—",       "ATTRIBUTION_WORKER_ENABLED=true", "Regex-first; AI only for ambiguous; 3K calls/day cap"),
    ]
    for proc in processes:
        for j, val in enumerate(proc):
            if j == 3:
                status_cell(ws, r, j + 1, val)
            else:
                data_cell(ws, r, j + 1, val,
                          bold=(j == 0),
                          wrap=(j in (2, 6)),
                          align="left" if j != 3 else "center")
        r += 1
    r += 1

    # ── Critical Numbers
    section_header(ws, r, 1, "  CRITICAL NUMBERS", span=COLS)
    r += 1
    num_headers = ["Metric", "Value", "As Of", "Notes"]
    for i, h in enumerate(num_headers):
        col_header(ws, r, i + 1, h)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
    r += 1
    numbers = [
        ("Messages ingested",           "147,793",           "June 2026", "Live + backfilled Telegram messages across all tracked groups"),
        ("DB migrations applied",        "34 / 34",           "June 2026", "All Phase 2 migrations complete"),
        ("Railway processes",            "8",                 "June 2026", "6 always-on + 2 gated"),
        ("Telegram accounts",            "3",                 "June 2026", "غيث (listener), راوي (backfill), إبراهيم (media)"),
        ("Course intelligence profiles", "339",               "June 2026", "Courses with any corpus coverage"),
        ("Exam intelligence signals",    "263",               "June 2026", "Recurring exam topics by course/type"),
        ("Message signals",              "3,179",             "June 2026", "exam_emphasis, difficulty, professor_note, resource_rec, confusion_cluster"),
        ("Est. monthly infra cost",      "~$50–80 USD",       "June 2026", "Railway + Supabase + OpenAI (variable)"),
        ("Target price (Phase 1)",       "SAR 79 / semester", "Hypothesis","Finals Companion tier — never charge for AI answers"),
    ]
    for num in numbers:
        for j, val in enumerate(num):
            if j == 3:
                c = ws.cell(row=r, column=4, value=val)
                c.font = Font(size=9, color="000000", italic=True)
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0),
                          align="center" if j == 1 else "left")
        r += 1
    r += 1

    # ── 10 Things to Know First
    section_header(ws, r, 1, "  10 THINGS A NEW PERSON MUST KNOW", span=COLS)
    r += 1
    things = [
        "1.  The corpus is the moat. 148K+ Telegram messages + official docs are the differentiated asset. AI is replaceable; the indexed community knowledge is not.",
        "2.  Three Telegram user accounts, one per worker type: غيث (listener), راوي (backfill), إبراهيم (media). NEVER share sessions — causes AuthKeyDuplicatedError.",
        "3.  All DB access is direct PostgREST over httpx. No ORM, no Supabase client library. See ADR-0009.",
        "4.  Backfill and live ingestion are permanently separated (ADR-0002). rumman_engine.py must never call iter_messages.",
        "5.  Intelligence and attribution workers are GATED. They cost money. Set env vars + review budgets before enabling.",
        "6.  All secrets live in Railway environment variables. The .env file is gitignored. Never commit credentials.",
        "7.  The claim model (machine_asserted → confirmed/rejected) is the anti-hallucination backbone. False attribution is worse than no attribution.",
        "8.  The search API has a selective LRU cache: only course-specific queries are cached (2hr TTL). Broad queries bypass cache to prevent cross-user data leakage.",
        "9.  Priority order for ingestion: Regulations → StudyPlans → CourseContent → OpenData. Use batch_ingest_seu.py for the 93 official SEU docs.",
        "10. The 'media' worker had 11K failed jobs due to a session env var mismatch (commit 3dd5332). After confirming TELEGRAM_MEDIA_IBRAHIM_SESSION is set in Railway, run reset_media_jobs.py.",
    ]
    for thing in things:
        c = ws.cell(row=r, column=1, value=f"  {thing}")
        c.font = Font(size=9, color="000000")
        c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _thin_bottom()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    # ── Emergency Actions
    section_header(ws, r, 1, "  IF SOMETHING BREAKS", span=COLS)
    r += 1
    em_headers = ["Situation", "First Check", "Action"]
    for i, h in enumerate(em_headers):
        col_header(ws, r, i + 1, h)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
    r += 1
    emergencies = [
        ("Bot not responding to students",   "Railway → bot process logs",           "Restart bot service. Check TELEGRAM_BOT_TOKEN. Verify /synthesize endpoint is up."),
        ("New messages not being ingested",  "Railway → listener process logs",       "Check TELEGRAM_LISTENER_GHAYTH_SESSION. Verify session not expired. Restart listener."),
        ("Media/audio jobs stuck",           "Railway → media process logs",          "Confirm TELEGRAM_MEDIA_IBRAHIM_SESSION is set. Run reset_media_jobs.py after fix."),
        ("OpenAI costs spiking",             "intelligence_worker logs (run tokens)", "Disable INTELLIGENCE_WORKER_ENABLED=false. Check INTELLIGENCE_MAX_TOKENS_PER_RUN."),
        ("Search returning wrong answers",   "search_api.py LRU cache",               "Clear cache (restart search service). Check corpus coverage with refresh_course_profiles.py."),
        ("DB query failing unexpectedly",    "Supabase SQL Editor → Table Editor",   "Check if migration needed. Verify table exists. Review match_documents() RPC."),
        ("Session expired (AuthKeyError)",   "Telegram auth failure in logs",         "Run auth_session.py locally for the affected account. Update Railway env var."),
        ("Backfill jobs stalled",            "telegram_backfill_jobs with status=running + expired lease", "Expired leases auto-reset. Manual fix: UPDATE status='pending' WHERE lease_expires_at < now()."),
    ]
    for em in emergencies:
        for j, val in enumerate(em):
            if j == 2:
                c = ws.cell(row=r, column=3, value=val)
                c.font = Font(size=9, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=True)
        ws.row_dimensions[r].height = 32
        r += 1

    r += 2

    # ── Quick Commands
    section_header(ws, r, 1, "  QUICK COMMANDS  (most-needed CLI ops — run from repo root with .env present)", span=COLS)
    r += 1
    col_header(ws, r, 1, "Command")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    col_header(ws, r, 4, "What It Does")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
    r += 1
    commands = [
        ("python3 scripts/reset_media_jobs.py",                 "Re-queue ~11K failed telegram_media jobs (after confirming TELEGRAM_MEDIA_IBRAHIM_SESSION set in Railway)"),
        ("python3 scripts/batch_ingest_seu.py --dry-run",       "Preview bulk-ingest of 93 official SEU docs. Remove --dry-run to execute."),
        ("python3 scripts/weekly_report.py",                    "Weekly ops health report → posts to Telegram ops channel (requires RUMMAN_OPS_CHAT_ID)"),
        ("python3 scripts/gap_analyst.py",                      "Cluster zero-result learning_events → gap_items table (run monthly)"),
        ("python3 scripts/refresh_course_profiles.py",          "Recompute course_intelligence_profiles + exam_intelligence coverage metrics"),
        ("python3 scripts/message_signal_worker.py",            "Extract typed signals from 148K message corpus → message_signals (already 3,179 signals; re-run to catch new messages)"),
        ("python3 scripts/eval_bot_quality.py",                 "10 test queries — synthesis quality baseline (run before charging students)"),
        ("python3 app/query_handler.py IT362 'exam topics'",    "Test query synthesis locally for a specific course (dev/debug only)"),
    ]
    for cmd in commands:
        c1 = ws.cell(row=r, column=1, value=f"  {cmd[0]}")
        c1.font = Font(name="Courier New", size=8, bold=True, color="000000")
        c1.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c1.alignment = _align("left", "center", wrap=False)
        c1.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c2 = ws.cell(row=r, column=4, value=f"  {cmd[1]}")
        c2.font = Font(size=9, color="000000")
        c2.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c2.alignment = _align("left", "center", wrap=True)
        c2.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 20
        r += 1

    # ── Column widths
    widths = [28, 36, 42, 10, 12, 32, 50, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 2: SYSTEMS ───────────────────────────────────────────────────────────

def build_systems(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["SYSTEMS"]
    ws.freeze_panes = "A3"

    COLS = 10
    r = 1

    title_cell(ws, r, 1, "  SYSTEMS — All External Services, Accounts & Costs",
               span=COLS, bg=C_NAVY, size=13)
    r += 1
    title_cell(ws, r, 1,
               "  One row per external service. Single source of truth for accounts, costs, and credential names.",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 1

    headers = ["Service", "Category", "Account / Login",
               "Plan / Tier", "Monthly Cost (USD)",
               "Status", "Critical?", "Credential Variable(s)", "URL / Location", "Notes"]
    for i, h in enumerate(headers):
        col_header(ws, r, i + 1, h)
    r += 1

    services = [
        # Service | Category | Account | Plan | Cost | Status | Critical | Cred Vars | URL | Notes
        ("Railway",              "Infrastructure",  "rumman.ops@gmail.com",  "Hobby (scales to Pro)", "$20–40",  "LIVE",   "YES", "RAILWAY_API_TOKEN",                                        "railway.app",           "8 processes. Cost = base + resource usage per process."),
        ("Supabase",             "Database",        "rumman.ops@gmail.com",  "Pro",                   "~$25",    "LIVE",   "YES", "SUPABASE_URL, SUPABASE_KEY",                               "supabase.com/dashboard","34 migrations applied. Ref: yriavgczteuirigsvedu"),
        ("Supabase Storage",     "Object Storage",  "(same as Supabase)",    "Included in Pro",       "Included","LIVE",   "YES", "(uses SUPABASE_KEY)",                                      "Bucket: rumman-content","Stores ingested PDFs (raw files deleted after extraction)"),
        ("OpenAI",               "AI / ML",         "rumman.ops@gmail.com",  "Pay-per-use",           "~$15–30", "LIVE",   "YES", "OPENAI_API_KEY",                                           "platform.openai.com",   "gpt-4o-mini + text-embedding-3-large. Cost varies with query volume."),
        ("Telegram — غيث",       "Comms (Listener)","Phone: +966582282200",  "Personal account",      "Free",    "LIVE",   "YES", "TELEGRAM_LISTENER_GHAYTH_SESSION, TELEGRAM_API_ID, TELEGRAM_API_HASH","t.me","Passive listener only. Never sends messages."),
        ("Telegram — راوي",      "Comms (Backfill)","Phone: +966590111167",  "Personal account",      "Free",    "LIVE",   "YES", "TELEGRAM_BACKFILL_RAWI_SESSION",                           "t.me",                  "Historical ingestion. Must be joined to all 102+ tracked groups."),
        ("Telegram — إبراهيم",   "Comms (Media)",   "Phone: +966560064766",  "Personal account",      "Free",    "LIVE",   "YES", "TELEGRAM_MEDIA_IBRAHIM_SESSION",                           "t.me",                  "Media/file download. Fallback: TELEGRAM_BAYAN_SESSION (legacy)."),
        ("Telegram Bot",         "Comms (Bot)",     "BotFather token",       "Free",                  "Free",    "LIVE",   "YES", "TELEGRAM_BOT_TOKEN",                                       "t.me/@RummanBot (verify)","Student-facing. Does NOT use a StringSession — uses BotFather token."),
        ("GitHub",               "Source Control",  "IbraSQ / RUMMANOPS",    "Free",                  "Free",    "LIVE",   "NO",  "—",                                                        "github.com/RUMMANOPS/rumman-core","Main branch. ADR-0003: docs in repo are source of truth."),
        ("Google Workspace",     "Email / Ops",     "rumman.ops@gmail.com",  "Free Gmail",            "Free",    "LIVE",   "NO",  "—",                                                        "gmail.com",             "Ops email for service registrations."),
        ("Domain (if any)",      "Web",             "—",                     "—",                     "—",       "PLANNED","NO",  "—",                                                        "—",                     "No public domain yet. Bot is the primary student surface."),
    ]
    for svc in services:
        for j, val in enumerate(svc):
            if j == 5:
                status_cell(ws, r, j + 1, val)
            elif j == 6:
                status_cell(ws, r, j + 1, val)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=(j in (7, 9)), align="left")
        r += 1

    r += 2
    section_header(ws, r, 1, "  ESTIMATED MONTHLY COST BREAKDOWN", span=COLS)
    r += 1
    cost_headers = ["Line Item", "Estimated Cost (USD)", "Variable?", "Scales With", "Notes"]
    for i, h in enumerate(cost_headers):
        col_header(ws, r, i + 1, h)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
    r += 1
    costs = [
        ("Railway (infrastructure)",        "$20–40",  "Yes",  "Number of processes + resource usage", "8 processes. Add processes = cost increases."),
        ("Supabase (DB + storage)",         "~$25",    "No",   "Storage size (Pro plan flat + overages)", "At current scale well within Pro limits."),
        ("OpenAI — gpt-4o-mini synthesis",  "~$8–20",  "Yes",  "Query volume × tokens per synthesis",  "$0.15/1M input + $0.60/1M output. ~1,500 tok/query."),
        ("OpenAI — text-embedding-3-large", "~$2–5",   "Yes",  "New document chunks per period",        "$0.13/1M tokens. Decreases as corpus stabilizes."),
        ("OpenAI — intelligence worker",    "~$1–10",  "Yes",  "Messages processed × 0.15/1M tokens",  "Gated. Off by default. Budget-capped at 200K tok/run."),
        ("OpenAI — attribution worker",     "~$1–5",   "Yes",  "Unattributed chunks × calls/day",       "Gated. 3K calls/day cap. Regex-first reduces cost."),
        ("Telegram accounts (3)",           "Free",    "No",   "—",                                     "Phone numbers owned. No recurring cost."),
        ("TOTAL (estimated)",               "~$57–105","Yes",  "Query volume + ingestion volume",       "Lower end = quiet period. Upper = active usage."),
    ]
    for cost in costs:
        bold_row = cost[0].startswith("TOTAL")
        for j, val in enumerate(cost):
            if j == 4:
                c = ws.cell(row=r, column=5, value=val)
                c.font = Font(size=9, color="000000", italic=True, bold=bold_row)
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=bold_row or (j == 0), align="left")
        r += 1

    widths = [22, 16, 20, 18, 12, 9, 10, 42, 28, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 3: PROCESSES ─────────────────────────────────────────────────────────

def build_processes(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["PROCESSES"]
    ws.freeze_panes = "A3"
    COLS = 9
    r = 1

    title_cell(ws, r, 1, "  PROCESSES — Runtime Architecture & Configuration",
               span=COLS, bg=C_NAVY, size=13)
    r += 1
    title_cell(ws, r, 1,
               "  Technical operating manual: how every process works, what it needs, and how they relate.",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 2

    # ── Railway Processes
    section_header(ws, r, 1, "  RAILWAY PROCFILE — 8 PROCESSES", span=COLS)
    r += 1
    h = ["Process (Procfile key)", "File", "Role (one sentence)", "Status",
         "Telegram Session", "Key Env Vars", "Job Types Handled", "Notes"]
    for i, hh in enumerate(h):
        col_header(ws, r, i + 1, hh)
    r += 1
    procs = [
        ("listener",     "app/rumman_engine.py",              "Receives live Telegram messages → messages table",            "LIVE",   "TELEGRAM_LISTENER_GHAYTH_SESSION",  "SUPABASE_URL, SUPABASE_KEY, TELEGRAM_API_ID/HASH",         "—",                                "ENABLE_BACKFILL=False guard must stay False"),
        ("backfill",     "app/telegram_backfill_worker.py",   "Claims backfill_jobs rows, ingests history with lease",       "LIVE",   "TELEGRAM_BACKFILL_RAWI_SESSION",    "SUPABASE_URL, SUPABASE_KEY, TELEGRAM_API_ID/HASH",         "—",                                "Lease-based; gap-fill job support"),
        ("media",        "app/telegram_download_worker.py",   "Downloads Telegram media; transcribes audio via Whisper",     "LIVE",   "TELEGRAM_MEDIA_IBRAHIM_SESSION",    "SUPABASE_URL, SUPABASE_KEY, OPENAI_API_KEY, TELEGRAM_*",   "telegram_media, audio_transcribe", "~11K failed jobs pending reset"),
        ("embed",        "app/embed_worker.py",               "Chunks text and embeds via text-embedding-3-large",           "LIVE",   "—",                                 "SUPABASE_URL, SUPABASE_KEY, OPENAI_API_KEY",               "embed_chunk",                      "Question-aware exam chunking; NFKC normalization"),
        ("search",       "app/search_api.py",                 "FastAPI search: query understanding → pgvector → synthesis",  "LIVE",   "—",                                 "SUPABASE_URL, SUPABASE_KEY, OPENAI_API_KEY, PORT",         "—",                                "LRU cache 2h; course-specific only"),
        ("bot",          "app/telegram_bot.py",               "Student-facing Telegram bot; calls /synthesize",              "LIVE",   "TELEGRAM_BOT_TOKEN",                "SUPABASE_URL, SUPABASE_KEY, SEARCH_API_URL, TELEGRAM_*",   "—",                                "6-turn conversation history; enrollment persistence"),
        ("intelligence", "app/intelligence_worker.py",        "Extracts assignments/deadlines from messages",                "GATED",  "—",                                 "OPENAI_API_KEY, INTELLIGENCE_WORKER_ENABLED=true",         "—",                                "Budget: 200K tokens/run. Enable in Railway."),
        ("attribution",  "app/attribution_worker.py",         "AI course attribution for untagged document_chunks",          "GATED",  "—",                                 "OPENAI_API_KEY, ATTRIBUTION_WORKER_ENABLED=true",          "—",                                "Regex-first. 3K API calls/day cap. High threshold 0.85."),
    ]
    for proc in procs:
        for j, val in enumerate(proc):
            if j == 3:
                status_cell(ws, r, j + 1, val)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=(j in (2, 5, 7)), align="left")
        ws.row_dimensions[r].height = 36
        r += 1
    r += 2

    # ── Telegram Session Accounts
    section_header(ws, r, 1, "  TELEGRAM SESSION ACCOUNTS  (3 accounts, 1 session each)", span=COLS)
    r += 1
    h2 = ["Account Name", "Identity (Arabic)", "Phone", "Role", "Session Variable", "Fallback Variable", "Used By Process", "Rule"]
    for i, hh in enumerate(h2):
        col_header(ws, r, i + 1, hh)
    r += 1
    accounts = [
        ("Ghayth",   "غيث",     "+966582282200", "Passive listener",   "TELEGRAM_LISTENER_GHAYTH_SESSION", "TELEGRAM_GHAYTH_SESSION (remove)", "listener",    "PASSIVE ONLY — never sends, never responds"),
        ("Rawi",     "راوي",    "+966590111167", "Historical backfill", "TELEGRAM_BACKFILL_RAWI_SESSION",   "—",                               "backfill",    "Must be member of all 46+ groups"),
        ("Ibrahim",  "إبراهيم", "+966560064766", "Media & files",      "TELEGRAM_MEDIA_IBRAHIM_SESSION",   "TELEGRAM_BAYAN_SESSION (remove)",  "media",       "Bot token is separate — not a StringSession"),
    ]
    for acc in accounts:
        for j, val in enumerate(acc):
            data_cell(ws, r, j + 1, val, bold=(j in (0, 4)), wrap=(j == 7), align="left")
        r += 1
    r += 2

    # ── Processing Job Types
    section_header(ws, r, 1, "  PROCESSING JOB TYPES  (processing_jobs table)", span=COLS)
    r += 1
    h3 = ["job_type", "Worker Process", "Source", "Destination", "Status", "Notes"]
    for i, hh in enumerate(h3):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    r += 1
    job_types = [
        ("pdf_extract",        "pdf_worker.py (on-demand)", "Supabase Storage",   "source_documents.extracted_text",  "ACTIVE", "Run manually when ingesting official docs"),
        ("embed_chunk",        "embed (Railway)",           "source_documents",   "document_chunks",                  "ACTIVE", "Auto-queued by pdf_worker after extraction"),
        ("telegram_media",     "media (Railway)",           "messages table",     "media_files",                      "ACTIVE", "~11K pending after session env var fix"),
        ("audio_transcribe",   "media (Railway)",           "media_files",        "media_files.transcription",        "ACTIVE", "Unified into media worker (avoids session conflict)"),
        ("telegram_gap_fill",  "backfill (Railway)",        "telegram_sync_state","messages table",                   "ACTIVE", "Auto-created when ID jump detected on live message"),
    ]
    for jt in job_types:
        for j, val in enumerate(jt):
            if j == 4:
                status_cell(ws, r, j + 1, val)
            elif j == 5:
                c = ws.cell(row=r, column=6, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), align="left")
        r += 1
    r += 2

    # ── Required Environment Variables
    section_header(ws, r, 1, "  REQUIRED ENVIRONMENT VARIABLES  (Railway env vars — no values stored here)", span=COLS)
    r += 1
    h4 = ["Variable Name", "What It Is", "Required By", "Where to Set", "Notes"]
    for i, hh in enumerate(h4):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
    r += 1
    env_vars = [
        ("TELEGRAM_API_ID",                   "Telegram app ID (numeric)",           "listener, backfill, media",    "Railway env",  "From my.telegram.org — same value for all 3 accounts"),
        ("TELEGRAM_API_HASH",                 "Telegram app hash (string)",          "listener, backfill, media",    "Railway env",  "From my.telegram.org — same value for all 3 accounts"),
        ("TELEGRAM_LISTENER_GHAYTH_SESSION",  "StringSession for غيث account",       "listener",                     "Railway env",  "Generated by auth_session.py locally — never commit"),
        ("TELEGRAM_BACKFILL_RAWI_SESSION",    "StringSession for راوي account",       "backfill",                     "Railway env",  "Generated by auth_session.py locally — never commit"),
        ("TELEGRAM_MEDIA_IBRAHIM_SESSION",    "StringSession for إبراهيم account",   "media",                        "Railway env",  "Canonical. Fallback: TELEGRAM_BAYAN_SESSION (legacy)"),
        ("TELEGRAM_BOT_TOKEN",                "BotFather token for student bot",     "bot",                          "Railway env",  "NOT a StringSession — separate identity entirely"),
        ("SUPABASE_URL",                      "Supabase project REST API URL",       "all workers",                  "Railway env",  "Format: https://<ref>.supabase.co"),
        ("SUPABASE_KEY",                      "Supabase service-role key",           "all workers",                  "Railway env",  "Full access — keep secret. Never commit."),
        ("OPENAI_API_KEY",                    "OpenAI API key",                      "embed, search, intelligence, attribution, bot","Railway env","Shared across all AI calls. Monitor spend."),
        ("SEU_TENANT_ID",                     "SEU tenant UUID",                     "all workers",                  "Railway env",  "00000000-0000-0000-0000-000000000001 (default)"),
        ("SEARCH_API_URL",                    "Internal URL of search service",      "bot",                          "Railway env",  "Railway internal URL. bot → search communication."),
        ("INTELLIGENCE_WORKER_ENABLED",       "Gate for intelligence worker",        "intelligence",                 "Railway env",  "Set to 'true' to activate. Off by default."),
        ("ATTRIBUTION_WORKER_ENABLED",        "Gate for attribution worker",         "attribution",                  "Railway env",  "Set to 'true' to activate. Off by default."),
        ("RUMMAN_USER_SALT",                  "HMAC salt for user ID hashing",       "bot, search",                  "Railway env",  "Privacy: user IDs are hashed before storage"),
        ("RUMMAN_OPS_CHAT_ID",                "Telegram chat_id for ops channel",   "weekly_report.py",             "Railway env",  "Used only for weekly ops reports"),
    ]
    for ev in env_vars:
        for j, val in enumerate(ev):
            if j == 4:
                c = ws.cell(row=r, column=5, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0),
                          wrap=(j in (1, 2)), align="left")
        r += 1

    r += 2

    # ── Scripts Inventory
    section_header(ws, r, 1, "  SCRIPTS INVENTORY  (run locally — require .env — not deployed on Railway)", span=COLS)
    r += 1
    col_header(ws, r, 1, "Script File")
    col_header(ws, r, 2, "Category")
    col_header(ws, r, 3, "Purpose")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    col_header(ws, r, 7, "Run When")
    col_header(ws, r, 8, "Cost?")
    col_header(ws, r, 9, "Est. Cost")
    r += 1
    scripts_inv = [
        ("ingest_document.py",          "Ingestion",   "Ingest one file: upload → source_documents → pdf_extract job → embed_chunk job",                            "On demand",          "YES",  "~$0.01"),
        ("batch_ingest_seu.py",         "Ingestion",   "Bulk-ingest all 93 SEU official docs in priority order (Regulations → StudyPlans → CourseContent)",          "Once / semester",    "YES",  "~$0.50"),
        ("seed_courses.py",             "Ingestion",   "Push course catalog (codes, names, prereqs) from inst_courses.json into inst_courses table",                  "Setup / update",     "OPT",  "Free (or ~$0.05 with --embed)"),
        ("create_backfill_jobs.py",     "Backfill",    "Create telegram_backfill_jobs rows for specified chat IDs — backfill worker picks them up",                   "Adding new groups",  "NO",   "Free"),
        ("reset_media_jobs.py",         "Maintenance", "Re-queue failed telegram_media jobs back to pending (run after session env var fix)",                         "After incident",     "NO",   "Free"),
        ("backfill_tenant_id_033.py",   "Maintenance", "Backfill missing tenant_id on all rows (companion to migration 033)",                                         "One-time",           "NO",   "Free"),
        ("backfill_course_codes.py",    "Maintenance", "Regex + LLM inference to fill null course_code on source_documents and their chunks",                        "On demand",          "YES",  "~$0.10"),
        ("gap_analyst.py",              "Analysis",    "Cluster zero-result learning_events into gap_items (signals corpus coverage holes)",                          "Monthly",            "YES",  "~$0.05"),
        ("refresh_course_profiles.py",  "Analysis",    "Recompute course_intelligence_profiles and exam_intelligence metrics per course",                             "After ingestion",    "NO",   "Free"),
        ("extract_exam_signals.py",     "Analysis",    "Extract recurring exam topics from message corpus → exam_intelligence table",                                 "Monthly",            "YES",  "~$0.10"),
        ("message_signal_worker.py",    "Analysis",    "Extract typed signals (exam_emphasis, difficulty, professor_note, resource_rec) from 148K messages",         "Once + after ingest","YES",  "~$0.35"),
        ("weekly_report.py",            "Reporting",   "Weekly ops + product health report → posts to Telegram ops channel (RUMMAN_OPS_CHAT_ID required)",           "Weekly",             "YES",  "~$0.05"),
        ("eval_bot_quality.py",         "QA",          "10 test queries; before/after synthesis quality comparison. Baseline before launch.",                         "Pre-launch / change","YES",  "~$0.05"),
        ("generate_seed_lexicon.py",    "Lexicon",     "Analyze document_chunks for non-standard terms → seed_candidates_<timestamp>.json (gitignored)",             "On demand",          "NO",   "Free"),
        ("review_candidates.py",        "Lexicon",     "Interactive review of lexicon candidates before adding to data/normalization_dict.json",                     "After generate",     "NO",   "Free"),
        ("export_group_links.py",       "Utility",     "Export Telegram group invite links for all tracked chats via إبراهيم session",                               "On demand",          "NO",   "Free"),
        ("extract_concepts.py",         "Utility",     "Extract academic concepts from document_chunks for future knowledge graph seeding",                           "On demand",          "YES",  "~$0.05"),
        ("generate_ops_workbook.py",    "Meta",        "Generate this workbook — RUMMAN_OPS_YYYY-MM-DD.xlsx (gitignored output)",                                    "Any time",           "NO",   "Free"),
    ]
    for s in scripts_inv:
        data_cell(ws, r, 1, s[0], bold=True, align="left")
        data_cell(ws, r, 2, s[1], bold=False, align="left")
        c = ws.cell(row=r, column=3, value=s[2])
        c.font = Font(size=9, color="000000")
        c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        data_cell(ws, r, 7, s[3], bold=False, align="center")
        status_cell(ws, r, 8, s[4])
        data_cell(ws, r, 9, s[5], bold=False, align="center")
        ws.row_dimensions[r].height = 26
        r += 1

    widths = [38, 16, 24, 14, 12, 44, 26, 10, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 4: CORPUS ────────────────────────────────────────────────────────────

def build_corpus(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["CORPUS"]
    ws.freeze_panes = "A3"
    COLS = 8
    r = 1

    title_cell(ws, r, 1, "  CORPUS — The Data Asset, Migrations & Knowledge Layer",
               span=COLS, bg=C_NAVY, size=13)
    r += 1
    title_cell(ws, r, 1,
               "  The knowledge corpus is the product. This sheet tracks what exists, what is applied, and what is missing.",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 2

    # ── Corpus metrics
    section_header(ws, r, 1, "  CORPUS METRICS  (as of June 2026)", span=COLS)
    r += 1
    h = ["Metric", "Value", "Source Table", "Notes"]
    for i, hh in enumerate(h):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
    r += 1
    metrics = [
        ("Total messages ingested",          "147,793",       "messages",                          "Live + backfilled Telegram messages — as of June 2026"),
        ("Telegram groups tracked",           "102",           "telegram_backfill_jobs",            "Total backfill_jobs rows; 20 completed, rest pending or running"),
        ("Course intelligence profiles",      "339",           "course_intelligence_profiles",      "Courses with any corpus coverage computed"),
        ("Exam intelligence signals",         "263",           "exam_intelligence",                 "Recurring exam topics by (course, exam_type)"),
        ("Message signals",                   "3,179",         "message_signals",                   "exam_emphasis, difficulty, professor_note, resource_rec, confusion_cluster"),
        ("Document chunks (embedded)",        "127,070",       "document_chunks",                   "Vector-embedded chunks available for retrieval — as of June 2026"),
        ("Source documents (all types)",      "11,391",        "source_documents",                  "All ingested files: student exams, official docs, Telegram exports"),
        ("Source docs (official types)",      "1,014",         "source_documents",                  "source_type IN (exam, study_plan, regulation, course_description) — incl. student-uploaded exams"),
        ("SEU institutional repo ingested",   "~9 of 93",      "source_documents",                  "93 files in knowledge repository. Run batch_ingest_seu.py to complete bulk ingest."),
        ("Intelligence items extracted",      "184",           "intelligence_items",                "Worker was gated — 184 from partial runs; enable INTELLIGENCE_WORKER_ENABLED for full corpus"),
        ("Knowledge gap items",               "34",            "gap_items",                         "Clustered zero-result events from gap_analyst.py"),
        ("Learning events logged",            "87",            "learning_events",                   "Search + synthesis events — low count confirms system not yet in wide use"),
        ("Student context records",           "TBD",           "student_context",                   "Enrolled courses per user — populated when students register courses with bot"),
    ]
    for m in metrics:
        for j, val in enumerate(m):
            if j == 3:
                c = ws.cell(row=r, column=4, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), align="left")
        r += 1
    r += 2

    # ── University Knowledge Repository
    section_header(ws, r, 1, "  UNIVERSITY KNOWLEDGE REPOSITORY  (official docs — outside this repo)", span=COLS)
    r += 1
    title_cell(ws, r, 1,
               "  Location: .../0-RUMMAN/0-Universities/1- Saudi Electronic University/   "
               "  Ingest all: python3 scripts/batch_ingest_seu.py",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 1
    h_repo = ["Directory", "Content Type", "File Count", "Ingested", "Status", "How to Ingest"]
    for i, hh in enumerate(h_repo):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    r += 1
    repo_dirs = [
        ("2. Regulations/",        "Exam rules, student guides, procedures (PDF)",                      "~20", "~3",   "PENDING", "batch_ingest_seu.py  or  ingest_document.py <file> --source-type regulation"),
        ("1. StudyPlans/",         "Official program study plans (PDF/DOCX) by college → dept → program", "~30","~3", "PENDING", "batch_ingest_seu.py  or  ingest_document.py <file> --source-type study_plan"),
        ("4. CourseContent/",      "Individual course syllabi (PDF) — ENGT program (34 files)",          "~34", "~1",  "PENDING", "batch_ingest_seu.py  or  ingest_document.py <file> --source-type course_description"),
        ("0. OpenData/",           "Enrollment stats, faculty data (PDF)",                               "~5",  "~1",  "PENDING", "ingest_document.py <file> --source-type regulation"),
        ("5. Diplomas/",           "Applied College diploma program plans",                               "~10", "~0",  "PENDING", "batch_ingest_seu.py  (lowest priority)"),
        ("3. AcademicCalendar/",   "Semester dates and windows",                                          "~5",  "~1",  "DONE",    "Seeded via migration 017 → academic_calendar table. No re-ingest needed."),
        ("_metadata/",             "knowledge_manifest.json, program_index.json",                         "2",   "—",   "N/A",     "Index files only — not ingested into document_chunks"),
        ("TOTAL",                  "93 official documents",                                               "~93", "~9",  "PENDING", "python3 scripts/batch_ingest_seu.py  (ingests all in one command)"),
    ]
    for rd in repo_dirs:
        is_total = rd[0] == "TOTAL"
        for j, val in enumerate(rd):
            if j == 4:
                status_cell(ws, r, j + 1, val)
            elif j == 5:
                c = ws.cell(row=r, column=6, value=val)
                c.font = Font(name="Courier New", size=8, color="000000", bold=is_total)
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(is_total or j == 0),
                          align="center" if j in (2, 3) else "left",
                          wrap=(j == 1))
        ws.row_dimensions[r].height = 24
        r += 1
    r += 2

    # ── Key Tables
    section_header(ws, r, 1, "  KEY TABLES  (PostgREST access: {SUPABASE_URL}/rest/v1/<table>)", span=COLS)
    r += 1
    h2 = ["Table", "Layer", "Primary Key / Unique Constraint", "Purpose"]
    for i, hh in enumerate(h2):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
    r += 1
    tables = [
        ("messages",                    "Data Spine",      "(platform_chat_id, platform_message_id)",      "All ingested Telegram messages — canonical record"),
        ("telegram_sync_state",         "Data Spine",      "platform_chat_id",                             "Newest/oldest IDs per chat; gap detection state"),
        ("telegram_backfill_jobs",      "Data Spine",      "platform_chat_id",                             "Lease-based historical ingestion work queue"),
        ("processing_jobs",             "Data Spine",      "id (UUID)",                                    "Generic async job queue for all worker types"),
        ("source_documents",            "Knowledge",       "content_hash",                                 "Uploaded/ingested files awaiting or post-extraction"),
        ("document_chunks",             "Knowledge",       "id (UUID)",                                    "Vector-embedded chunks — the retrieval corpus"),
        ("media_files",                 "Knowledge",       "id (UUID)",                                    "Audio transcription results; media metadata"),
        ("inst_colleges",               "Institutional",   "id (UUID)",                                    "SEU college master data; telegram_chat_ids mapping"),
        ("inst_specializations",        "Institutional",   "id (UUID)",                                    "Program specializations; college linkage"),
        ("inst_courses",                "Institutional",   "(tenant_id, code)",                            "Course catalog: credit hours, levels, prereqs"),
        ("intelligence_items",          "Intelligence",    "(tenant_id, source_platform, source_message_id, item_type)", "Extracted assignments/deadlines (gated worker)"),
        ("extracted_items",             "Intelligence",    "id (UUID)",                                    "daily_brief output; tasks, deadlines, decisions"),
        ("learning_events",             "Observability",   "id (UUID)",                                    "Every search, synthesis, zero_result event — analytics"),
        ("ai_runs",                     "Observability",   "id (UUID)",                                    "AI call provenance trail — cost, tokens, subject"),
        ("worker_heartbeats",           "Observability",   "worker_id",                                    "Worker liveness — last beat, status, metadata"),
        ("student_context",             "Personalization", "(tenant_id, user_hash)",                       "Enrolled courses per student (persistent across sessions)"),
        ("analysis_runs",               "Analysis",        "id (UUID)",                                    "gap_analyst, qa_miner output log — append-only"),
        ("gap_items",                   "Analysis",        "id (UUID)",                                    "Normalised knowledge gap rows from gap_analyst"),
        ("message_signals",             "Analysis",        "(course_code, chat_name, signal_type, content_hash)", "Typed intelligence signals from message corpus"),
        ("course_intelligence_profiles","Analysis",        "(course_code, tenant_id)",                     "Computed coverage metrics by course — refreshed on demand"),
        ("exam_intelligence",           "Analysis",        "(course_code, tenant_id, exam_type)",          "Top recurring exam topics extracted by extract_exam_signals.py"),
        ("worker_cursors",              "Operations",      "worker_id",                                    "Cursor-tracked position for intelligence_worker"),
        ("academic_calendar",           "Institutional",   "id (UUID)",                                    "1447H Hijri semester dates; injected into synthesis context"),
        ("active_extracted_items",      "View",            "—",                                            "extracted_items filtered by temporal validity + not rejected"),
        ("active_document_chunks",      "View",            "—",                                            "document_chunks filtered by superseded_by IS NULL"),
    ]
    for t in tables:
        for j, val in enumerate(t):
            if j == 3:
                c = ws.cell(row=r, column=4, value=val)
                c.font = Font(size=9, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0),
                          align="left", wrap=(j == 2))
        r += 1
    r += 2

    # ── Migration Registry
    section_header(ws, r, 1, "  MIGRATION REGISTRY  (34 migrations — all applied)", span=COLS)
    r += 1
    col_header(ws, r, 1, "#")
    col_header(ws, r, 2, "File")
    col_header(ws, r, 3, "Tables / Columns Created or Modified")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    col_header(ws, r, 5, "Required By")
    col_header(ws, r, 6, "Status")
    r += 1
    migrations = [
        ("001", "001_daily_brief_tables.sql",                      "brief_runs, extracted_items, ai_runs",                                        "app/daily_brief.py",                     "APPLIED"),
        ("002", "002_processing_jobs_retry_count.sql",             "processing_jobs.retry_count",                                                 "app/telegram_download_worker.py",        "APPLIED"),
        ("003", "003_knowledge_layer.sql",                         "source_documents, document_chunks (VECTOR 1536), pgvector",                   "app/embed_worker.py",                    "APPLIED"),
        ("004", "004_media_lifecycle.sql",                         "media_files",                                                                 "app/telegram_download_worker.py",        "APPLIED"),
        ("005", "005_match_documents_rpc.sql",                     "match_documents() pgvector RPC",                                              "app/search_api.py",                      "APPLIED"),
        ("006", "006_query_intelligence.sql",                      "query_logs, feedback — DROPPED in 016. Do not reference.",                    "superseded",                             "APPLIED"),
        ("007", "007_platform_foundations.sql",                    "tenants, users, sessions, events",                                            "all workers",                            "APPLIED"),
        ("008", "008_curriculum_foundations.sql",                  "seu_colleges, seu_specializations, seu_courses (renamed inst_* in 014)",       "scripts/seed_courses.py",                "APPLIED"),
        ("009", "009_curriculum_graduate_and_remapping.sql",       "8 graduate specializations, course remapping",                                "scripts/seed_courses.py",                "APPLIED"),
        ("010", "010_source_authority.sql",                        "document_chunks.source_authority tier column",                                "embed_worker, search_api",               "APPLIED"),
        ("011", "011_intelligence_layer.sql",                      "intelligence_items",                                                          "app/intelligence_worker.py",             "APPLIED"),
        ("012", "012_messages_tenant_id.sql",                      "messages.tenant_id (backfill)",                                               "app/rumman_engine.py",                   "APPLIED"),
        ("013", "013_embedding_model.sql",                         "document_chunks.embedding_model",                                             "app/embed_worker.py",                    "APPLIED"),
        ("014", "014_rename_seu_to_inst.sql",                      "seu_* → inst_* table renames",                                                "all institutional queries",              "APPLIED"),
        ("015", "015_claim_model_and_authority.sql",               "machine_asserted, confidence_tier columns",                                   "app/attribution_worker.py",              "APPLIED"),
        ("016", "016_temporal_and_ops.sql",                        "learning_events (new); query_logs + feedback DROPPED",                        "app/search_api.py",                      "APPLIED"),
        ("017", "017_academic_calendar_1447h.sql",                 "academic_calendar table; 1447H dates seeded",                                 "app/search_api.py",                      "APPLIED"),
        ("018", "018_drop_legacy_courses_table.sql",               "Drop legacy courses table",                                                   "—",                                      "APPLIED"),
        ("019", "019_fix_intelligence_items.sql",                  "Fix dedup constraint on intelligence_items",                                  "app/intelligence_worker.py",             "APPLIED"),
        ("020", "020_drop_seu_compat_views.sql",                   "Drop backward-compat seu_* views",                                            "—",                                      "APPLIED"),
        ("021", "021_ai_runs_defaults.sql",                        "Defaults/constraints on ai_runs",                                             "all AI workers",                         "APPLIED"),
        ("022", "022_match_documents_authority_tier.sql",          "Update match_documents() with authority tier filter",                         "app/search_api.py",                      "APPLIED"),
        ("023", "023_worker_heartbeats.sql",                       "worker_heartbeats liveness table",                                            "all workers",                            "APPLIED"),
        ("024", "024_course_names_bulk.sql",                       "Bulk Arabic course names in inst_courses",                                    "—",                                      "APPLIED"),
        ("025", "025_claim_model_temporal_and_contradiction.sql",  "active_extracted_items view, active_document_chunks view, supersession cols", "app/search_api.py",                      "APPLIED"),
        ("026", "026_analysis_runs.sql",                           "analysis_runs, gap_items",                                                    "scripts/gap_analyst.py",                 "APPLIED"),
        ("027", "027_document_chunks_metadata.sql",                "document_chunks.metadata JSONB column",                                       "embed_worker, search_api",               "APPLIED"),
        ("028", "028_self_healing_ingestion.sql",                  "Self-healing backfill improvements",                                          "app/telegram_backfill_worker.py",        "APPLIED"),
        ("029", "029_fix_intelligence_items_dedup.sql",            "Fix dedup logic on intelligence_items",                                       "app/intelligence_worker.py",             "APPLIED"),
        ("030", "030_student_context.sql",                         "student_context persistent cross-session memory",                             "app/search_api.py",                      "APPLIED"),
        ("031", "031_course_intelligence_profiles.sql",            "course_intelligence_profiles, exam_intelligence",                             "app/search_api.py",                      "APPLIED"),
        ("032", "032_message_signals.sql",                         "message_signals typed signals",                                               "scripts/message_signal_worker.py",       "APPLIED"),
        ("033", "033_backfill_tenant_id.sql",                      "Backfill missing tenant_id values",                                           "—",                                      "APPLIED"),
        ("034", "034_match_documents_fix.sql",                     "Fix match_documents(): filter_tenant UUID param, metadata JSONB return",      "app/search_api.py",                      "APPLIED"),
    ]
    for mig in migrations:
        for j, val in enumerate(mig):
            if j == 4:
                status_cell(ws, r, 6, val)
            elif j == 3:
                data_cell(ws, r, 5, val, align="left")
            elif j == 2:
                c = ws.cell(row=r, column=3, value=val)
                c.font = Font(size=9, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0),
                          align="center" if j == 0 else "left")
        r += 1

    widths = [32, 48, 42, 26, 34, 12, 8, 8]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 5: STRATEGY ─────────────────────────────────────────────────────────

def build_strategy(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["STRATEGY"]
    ws.freeze_panes = "A3"
    COLS = 8
    r = 1

    title_cell(ws, r, 1, "  STRATEGY — Product Direction, Stages & Commercial Model",
               span=COLS, bg=C_NAVY, size=13)
    r += 1
    title_cell(ws, r, 1,
               "  Where RUMMAN is going, why, and how it will sustain itself.",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 2

    # ── Product Identity
    section_header(ws, r, 1, "  PRODUCT IDENTITY  (what we are building and why)", span=COLS)
    r += 1
    identity = [
        ("Core thesis",          "Community knowledge, accumulated and organized, is the product. The AI is the retrieval and synthesis lens. The corpus is the moat."),
        ("Who it serves",        "Saudi university students facing information asymmetry: unofficial knowledge is scattered across 102+ Telegram groups, official docs are inaccessible."),
        ("The non-obvious bet",  "Students are not searching for an AI tutor. They want confidence before exams. Accumulated community intelligence (what the professor actually emphasizes, what appears in past exams) is what they will pay for."),
        ("What we are NOT",      "Not a search engine. Not an LLM wrapper. Not a tutoring service. Not a competitor to professors. Not a content creator."),
        ("The defensibility",    "A competitor can copy the AI. They cannot copy 148K indexed messages from 102+ student groups + 6 semesters of exam signals. The corpus is non-replicable."),
        ("The north star",       "Every student at every Saudi university has a trusted academic companion that knows their courses, their professors, and their exam patterns."),
    ]
    for item in identity:
        data_cell(ws, r, 1, item[0], bold=True, align="left")
        c = ws.cell(row=r, column=2, value=item[1])
        c.font = Font(size=9, color="000000")
        c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 36
        r += 1
    r += 2

    # ── Product Stages
    section_header(ws, r, 1, "  PRODUCT STAGES  (evolution path)", span=COLS)
    r += 1
    h = ["Stage", "Name", "Core Value Proposition", "Status", "Unlock Condition", "Key Metrics", "Notes"]
    for i, hh in enumerate(h):
        col_header(ws, r, i + 1, hh)
    r += 1
    stages = [
        ("Stage 1", "Finals Companion",     "72-hour exam preparation assistant. Knows what the professor emphasizes, what past exams cover, and what the study plan requires. 148K messages indexed.",   "CURRENT", "Bot live + corpus growing",              "Zero-result rate, query volume, corpus coverage by course", "One semester, one university, high-frequency use case"),
        ("Stage 2", "Semester Companion",   "Full semester intelligence: assignment tracking, deadlines, announcements from all groups in one place.",                              "PLANNED", "Intelligence worker active + enough data", "Active weekly users, engagement depth",                     "Requires intelligence_items to be live and populated"),
        ("Stage 3", "Academic OS",          "Cross-course, cross-semester intelligence. Course selection, prerequisite mapping, grade optimization.",                              "PLANNED", "Stage 2 proven + expanded corpus",         "Retention, word-of-mouth coefficient",                      "Requires institutional data depth + validated trust"),
        ("Stage 4", "Multi-University",     "Expand to other Saudi universities (KFUPM, KSU, KAUST, private). Per-tenant corpus, shared infrastructure.",                         "PLANNED", "Stage 3 validated + tenant architecture",   "Universities onboarded, cross-tenant query patterns",       "Multi-tenant architecture already in codebase (tenant_id)"),
    ]
    for stage in stages:
        for j, val in enumerate(stage):
            if j == 3:
                status_cell(ws, r, j + 1, val)
            else:
                data_cell(ws, r, j + 1, val, bold=(j in (0, 1)),
                          wrap=(j in (2, 4, 5, 6)), align="left")
        ws.row_dimensions[r].height = 48
        r += 1
    r += 2

    # ── Phase Roadmap
    section_header(ws, r, 1, "  PHASE ROADMAP  (technical phases)", span=COLS)
    r += 1
    col_header(ws, r, 1, "Phase")
    col_header(ws, r, 2, "Focus")
    col_header(ws, r, 3, "Status")
    col_header(ws, r, 4, "What Was Built")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    col_header(ws, r, 6, "Remaining")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    r += 1
    phases = [
        ("Phase 1", "Data spine + live ingestion",     "DONE",    "Listener, backfill, embed workers. message/chunk schema. 147,793 messages, 127,070 chunks.",                  "—"),
        ("Phase 2", "Intelligence + synthesis layer",  "CURRENT", "search_api (FastAPI), telegram bot, intelligence_items, claim model, calendar injection, gap_analyst, QA mining, message signals, student context, course profiles.", "Enable intelligence worker. Ingest 93 official SEU docs. Populate college chat_id mapping."),
        ("Phase 3", "Personalization + scale",         "PLANNED", "—",                                                                                        "Cross-session memory depth, recommendation engine, multi-program coverage"),
        ("Phase 4", "Multi-university expansion",      "PLANNED", "—",                                                                                        "Tenant onboarding flow, per-university corpus isolation, billing"),
    ]
    for phase in phases:
        data_cell(ws, r, 1, phase[0], bold=True, align="center")
        data_cell(ws, r, 2, phase[1], bold=True, align="left")
        status_cell(ws, r, 3, phase[2])
        for col_off, val in enumerate([phase[3], phase[4]]):
            col_s = 4 + col_off * 2
            col_e = col_s + 1 if col_off == 0 else COLS
            c = ws.cell(row=r, column=col_s, value=val)
            c.font = Font(size=9, color="000000")
            c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
            c.alignment = _align("left", "center", wrap=True)
            c.border = _border(style="hair")
            ws.merge_cells(start_row=r, start_column=col_s,
                           end_row=r, end_column=col_e)
        ws.row_dimensions[r].height = 54
        r += 1
    r += 2

    # ── Monetization Model
    section_header(ws, r, 1, "  MONETIZATION MODEL  (hypothesis — not validated)", span=COLS)
    r += 1
    h3 = ["Item", "Value / Principle", "Rationale"]
    for i, hh in enumerate(h3):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
    r += 1
    monetization = [
        ("Target price",          "SAR 79 / semester",             "Equals one textbook. Low enough to be impulsive. High enough to be sustainable."),
        ("What students pay for", "Access to accumulated intelligence", "NOT AI answers. NOT a chatbot. Access to indexed community knowledge that took years to accumulate."),
        ("Free tier philosophy",  "Search and basic synthesis free", "Free users contribute usage data (learning_events) that makes the paid product better."),
        ("What is NEVER charged", "AI answers, query responses",    "Charging per AI call commoditizes the product. The moat is the corpus, not the model."),
        ("Revenue at scale",      "~SAR 790K at 10K students",     "10,000 students × SAR 79 = SAR 790K/semester. Infra cost is ~$500/month at that scale."),
        ("Cost structure",        "Infra scales linearly, not exponentially", "Railway + Supabase + OpenAI. At 10K users, est. $300-600/mo. Margin > 95%."),
        ("B2B path (Phase 4)",    "University institutional licensing", "Sell to university libraries/support offices. Higher ACV, lower CAC."),
    ]
    for m in monetization:
        data_cell(ws, r, 1, m[0], bold=True, align="left")
        data_cell(ws, r, 2, m[1], bold=False, align="left")
        c = ws.cell(row=r, column=3, value=m[2])
        c.font = Font(size=9, italic=True, color="000000")
        c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 28
        r += 1
    r += 2

    # ── Key Metrics to Track
    section_header(ws, r, 1, "  KEY METRICS TO TRACK  (what matters at each stage)", span=COLS)
    r += 1
    h4 = ["Metric", "Stage", "How to Measure", "Target (Stage 1)"]
    for i, hh in enumerate(h4):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    r += 1
    key_metrics = [
        ("Zero-result rate",            "All",    "learning_events WHERE event_type='zero_result' / total queries",  "< 20% (corpus coverage drives this down)"),
        ("P95 synthesis latency",       "All",    "learning_events.latency_ms percentile",                          "< 8 seconds"),
        ("Weekly active users (WAU)",   "All",    "COUNT(DISTINCT user_hash) WHERE occurred_at > 7 days ago",       "Track; no absolute target yet"),
        ("Corpus coverage by course",   "Stage 1","course_intelligence_profiles: strong + moderate count",          "30+ courses with 'moderate' or 'strong' coverage"),
        ("Exam signal density",         "Stage 1","exam_intelligence: courses with top_topics populated",           "80+ courses with at least 'medium' confidence signals"),
        ("Bot response acceptance rate","Stage 1","User does NOT send follow-up clarification request",              "Proxy metric — needs explicit feedback to measure"),
        ("Message signals per course",  "Stage 2","message_signals grouped by course_code",                         "500+ signals before enabling intelligence as primary source"),
        ("Intelligence item precision", "Stage 2","Manually verified sample of intelligence_items",                  "Confidence ≥ 0.85 → real item in 90%+ of spot checks"),
    ]
    for km in key_metrics:
        data_cell(ws, r, 1, km[0], bold=True, align="left")
        data_cell(ws, r, 2, km[1], bold=False, align="center")
        c2 = ws.cell(row=r, column=3, value=km[2])
        c2.font = Font(size=9, color="000000")
        c2.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c2.alignment = _align("left", "center", wrap=True)
        c2.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        c3 = ws.cell(row=r, column=6, value=km[3])
        c3.font = Font(size=9, italic=True, color="000000")
        c3.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c3.alignment = _align("left", "center", wrap=True)
        c3.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 28
        r += 1

    widths = [28, 16, 48, 36, 16, 38, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 6: OPEN ITEMS ────────────────────────────────────────────────────────

def build_open_items(ws):
    ws.sheet_properties.tabColor = TAB_COLORS["OPEN ITEMS"]
    ws.freeze_panes = "A3"
    COLS = 8
    r = 1

    title_cell(ws, r, 1, "  OPEN ITEMS — Risks, Decisions, Next Actions & Known Debt",
               span=COLS, bg=C_NAVY, size=13)
    r += 1
    title_cell(ws, r, 1,
               "  Everything unresolved. Review at each work session. Items here should shrink over time.",
               span=COLS, bg=C_DARK, size=9, bold=False)
    r += 2

    # ── Active Risks
    section_header(ws, r, 1, "  ACTIVE RISKS", span=COLS)
    r += 1
    h = ["Risk", "Category", "Probability", "Impact", "Status", "Mitigation"]
    for i, hh in enumerate(h):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    r += 1
    risks = [
        ("Telegram account suspension (غيث/راوي/إبراهيم)",    "Operational", "MEDIUM", "HIGH",   "OPEN",   "Minimize API calls. Avoid flood limits. Keep accounts active (send periodic messages from personal phone)."),
        ("OpenAI API key compromised",                        "Security",    "LOW",    "HIGH",   "OPEN",   "Rotate key immediately via OpenAI dashboard. Update Railway env var. No retry logic — fast fail is correct."),
        ("Corpus quality degrades silently",                  "Quality",     "MEDIUM", "HIGH",   "OPEN",   "Monitor zero-result rate weekly via weekly_report.py. Run gap_analyst.py monthly."),
        ("Intelligence worker runaway cost",                  "Financial",   "MEDIUM", "MEDIUM", "OPEN",   "Hard-capped at MAX_TOKENS_PER_RUN (200K default). Gated by env var. Monitor via ai_runs table."),
        ("Supabase vector search latency at scale",           "Technical",   "LOW",    "MEDIUM", "OPEN",   "HNSW index in place. Fallback: add similarity threshold. Parallel search for enrolled courses."),
        ("Session key rotation not automated",                "Operational", "LOW",    "MEDIUM", "OPEN",   "Sessions can expire. auth_session.py is local-only. Update Railway env var manually after rotation."),
        ("Wrong course attribution (false positive)",         "Quality",     "MEDIUM", "HIGH",   "OPEN",   "Attribution worker uses 0.85 confidence threshold (intentionally high). Regex-first avoids AI for explicit codes."),
        ("Student data privacy (user hashes)",                "Legal",       "LOW",    "HIGH",   "OPEN",   "User IDs hashed via RUMMAN_USER_SALT before storage. Message text not linked to identifiable users in analytics."),
        ("Media worker 11K failed jobs backlog",              "Operational", "HIGH",   "MEDIUM", "OPEN",   "Confirm TELEGRAM_MEDIA_IBRAHIM_SESSION set in Railway. Run reset_media_jobs.py. ~15-20h drain time."),
        ("College chat_id mapping empty",                     "Data",        "HIGH",   "MEDIUM", "OPEN",   "inst_colleges.telegram_chat_ids not populated. Run populate script or manually update via Supabase."),
    ]
    for risk in risks:
        for j, val in enumerate(risk):
            if j in (2, 3, 4):
                status_cell(ws, r, j + 1, val)
            elif j == 5:
                c = ws.cell(row=r, column=6, value=val)
                c.font = Font(size=9, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=(j == 0), align="left")
        ws.row_dimensions[r].height = 32
        r += 1
    r += 2

    # ── Open Decisions
    section_header(ws, r, 1, "  OPEN DECISIONS  (unresolved strategic and technical questions)", span=COLS)
    r += 1
    h2 = ["Decision / Question", "Category", "Priority", "Status", "Options / Notes"]
    for i, hh in enumerate(h2):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
    r += 1
    decisions = [
        ("When do we launch to paying students? What is the MVP threshold?",            "Product",    "HIGH",   "OPEN", "Possible trigger: 30+ courses with strong/moderate coverage AND zero-result rate < 20%."),
        ("What is the onboarding flow for new students?",                               "Product",    "HIGH",   "OPEN", "Bot-first or web landing page? How does a student 'sign up'?"),
        ("How do we handle contradictory community vs. official information?",           "Technical",  "HIGH",   "OPEN", "Claim model exists. Synthesis prompt handles it. But no student-facing 'this is contested' UX yet."),
        ("Should we charge per-query, per-semester, or per-course?",                    "Commercial", "HIGH",   "OPEN", "Hypothesis: SAR 79/semester. Not validated. Could be per-course bundle (SAR 29/course)."),
        ("Which additional universities to target first after SEU?",                    "Strategic",  "MEDIUM", "OPEN", "KFUPM, KSU, or private universities? Depends on Telegram group density and student willingness."),
        ("Should the bot send proactive notifications (upcoming deadlines)?",           "Product",    "MEDIUM", "OPEN", "Requires student opt-in. Would dramatically increase engagement. Needs intelligence worker live."),
        ("How to handle Arabic dialect diversity across regions?",                      "Technical",  "MEDIUM", "OPEN", "normalization_dict.json covers Gulf Arabic. Needs expansion for non-Gulf dialects at multi-university scale."),
        ("Should RUMMAN become an agent (takes actions) or stay a reference?",          "Strategic",  "LOW",    "OPEN", "Current: reference only. Agent path (registers for exams, submits assignments) has liability risk."),
        ("What happens to the corpus if Telegram groups become private or deleted?",    "Risk",       "MEDIUM", "OPEN", "Messages already ingested are preserved. New messages lost. Backup strategy needed."),
        ("How do we validate corpus quality before charging?",                          "Product",    "HIGH",   "OPEN", "eval_bot_quality.py exists. Needs systematic test queries per course. Define 'good enough'."),
    ]
    for dec in decisions:
        for j, val in enumerate(dec):
            if j in (2, 3):
                status_cell(ws, r, j + 1, val)
            elif j == 4:
                c = ws.cell(row=r, column=5, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=(j == 0), align="left")
        ws.row_dimensions[r].height = 36
        r += 1
    r += 2

    # ── Next Priority Actions
    section_header(ws, r, 1, "  NEXT PRIORITY ACTIONS  (ordered by impact)", span=COLS)
    r += 1
    h3 = ["#", "Action", "Owner", "Status", "How to Do It"]
    for i, hh in enumerate(h3):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
    r += 1
    actions = [
        ("1", "Confirm TELEGRAM_MEDIA_IBRAHIM_SESSION is set in Railway; run reset_media_jobs.py",  "Ibrahim", "OPEN",   "Check Railway env vars → confirm var exists → python3 scripts/reset_media_jobs.py"),
        ("2", "Bulk-ingest 93 official SEU documents",                                               "Ibrahim", "OPEN",   "python3 scripts/batch_ingest_seu.py → python3 app/pdf_worker.py → embed_worker picks up"),
        ("3", "Populate inst_colleges.telegram_chat_ids",                                            "Ibrahim", "OPEN",   "SQL: UPDATE inst_colleges SET telegram_chat_ids='{chat_id,...}' WHERE name='...'"),
        ("4", "Enable intelligence worker (INTELLIGENCE_WORKER_ENABLED=true in Railway)",            "Ibrahim", "OPEN",   "Set env var in Railway → verify 011_intelligence_layer.sql applied → monitor ai_runs cost"),
        ("5", "Re-run message_signal_worker.py to catch new messages (3,179 signals exist; 148K corpus)",  "Ibrahim", "OPEN",   "python3 scripts/message_signal_worker.py (incremental run on new messages since last run)"),
        ("6", "Run extract_exam_signals.py and refresh_course_profiles.py",                          "Ibrahim", "OPEN",   "python3 scripts/extract_exam_signals.py && python3 scripts/refresh_course_profiles.py"),
        ("7", "Run eval_bot_quality.py to baseline synthesis quality",                               "Ibrahim", "OPEN",   "python3 scripts/eval_bot_quality.py → document before/after for launch readiness"),
        ("8", "Define and document the student onboarding flow",                                     "Ibrahim", "OPEN",   "Decide: bot-first or web? Write the flow. Implement /help and welcome message improvements."),
        ("9", "Set up weekly_report.py as a scheduled cron (Railway Cron Job)",                      "Ibrahim", "OPEN",   "Add cron trigger in Railway. Requires RUMMAN_OPS_CHAT_ID env var set."),
        ("10","Validate product with 5-10 real students before charging",                            "Ibrahim", "OPEN",   "Give 10 students free access. Collect qualitative feedback. Measure zero-result rate on real queries."),
    ]
    for act in actions:
        for j, val in enumerate(act):
            if j == 3:
                status_cell(ws, r, j + 1, val)
            elif j == 4:
                c = ws.cell(row=r, column=5, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j in (0, 1)),
                          wrap=(j in (1, 4)), align="left" if j != 0 else "center")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 2

    # ── Known Technical Debt
    section_header(ws, r, 1, "  KNOWN TECHNICAL DEBT  (non-blocking but should be resolved)", span=COLS)
    r += 1
    h4 = ["Item", "Severity", "File", "Notes"]
    for i, hh in enumerate(h4):
        col_header(ws, r, i + 1, hh)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
    r += 1
    debt = [
        ("TELEGRAM_BAYAN_SESSION legacy fallback — remove from code once Railway is confirmed to use IBRAHIM_SESSION",  "MEDIUM", "telegram_download_worker.py",   "Both vars exist as fallback chain. Clean up once confirmed."),
        ("TELEGRAM_GHAYTH_SESSION and TELEGRAM_SESSION_STRING legacy fallbacks in listener",                            "LOW",    "rumman_engine.py",              "Three-variable fallback chain. Remove old vars from Railway, then clean code."),
        ("pdf_worker.py not in Procfile — ingestion requires manual local run",                                         "LOW",    "Procfile + pdf_worker.py",      "Fine for current volume. Add to Railway if ingestion becomes frequent."),
        ("no_conflict handling on create_extract_job in batch_ingest_seu.py",                                           "LOW",    "scripts/batch_ingest_seu.py",   "Re-running batch will create duplicate pdf_extract jobs. Add on_conflict handling."),
        ("SUPABASE_REF hardcoded in refresh_course_profiles.py",                                                        "LOW",    "scripts/refresh_course_profiles.py","Project ref yriavgczteuirigsvedu. Not a secret but should be env var."),
        ("canonical_key() in export_group_links.py is incorrect for short chat IDs",                                    "LOW",    "scripts/export_group_links.py", "Function works in practice but arithmetic breaks for short IDs. Use _norm_chat_id() instead."),
        ("daily_brief.py fetch_messages lacks tenant_id filter",                                                        "LOW",    "app/daily_brief.py",            "Local tool only. Not deployed. Fine for single-tenant. Fix before multi-tenant."),
        ("No automated migration runner — all migrations are manual via Supabase SQL Editor",                           "MEDIUM", "supabase/migrations/",          "ADR-0003 compliance. See supabase/README.md for migration procedure."),
    ]
    for d in debt:
        for j, val in enumerate(d):
            if j == 1:
                status_cell(ws, r, j + 1, val)
            elif j == 3:
                c = ws.cell(row=r, column=4, value=val)
                c.font = Font(size=9, italic=True, color="000000")
                c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
                c.alignment = _align("left", "center", wrap=True)
                c.border = _border(style="hair")
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=COLS)
            else:
                data_cell(ws, r, j + 1, val, bold=(j == 0), wrap=(j in (0, 2)), align="left")
        ws.row_dimensions[r].height = 30
        r += 1

    widths = [50, 14, 34, 14, 10, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Sheet 7: KEYS ─────────────────────────────────────────────────────────────

def build_credentials(ws, env: dict):
    ws.sheet_properties.tabColor = TAB_COLORS["KEYS"]
    ws.freeze_panes = "A5"
    COLS = 7
    r = 1

    title_cell(ws, r, 1,
               "  ⚠  KEYS — API Keys, Tokens & Session Strings   "
               "  SENSITIVE: do not share, screenshot, or email this file.",
               span=COLS, bg="922B21", size=12)
    r += 1
    title_cell(ws, r, 1,
               f"  Generated: {date.today().isoformat()}   ·   "
               "Values read from local .env at generation time.   "
               "Variables not in .env are Railway-only — retrieve from Railway dashboard.",
               span=COLS, bg=C_CRED_DARK, size=9, bold=False)
    r += 2

    # ── Active credentials table
    section_header(ws, r, 1, "  ALL ACTIVE CREDENTIALS", span=COLS)
    r += 1
    col_header(ws, r, 1, "Variable Name")
    col_header(ws, r, 2, "Category")
    col_header(ws, r, 3, "Used By")
    col_header(ws, r, 4, "Source")
    col_header(ws, r, 5, "Current Value")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
    r += 1

    _NA = "NOT IN .env — retrieve from Railway dashboard (Settings → Variables)"

    creds = [
        # var_name | category | used_by | source_label | env_key (or literal)
        ("TELEGRAM_API_ID",                  "Telegram Auth",    "listener, backfill, media",           ".env",    "TELEGRAM_API_ID"),
        ("TELEGRAM_API_HASH",                "Telegram Auth",    "listener, backfill, media",           ".env",    "TELEGRAM_API_HASH"),
        ("TELEGRAM_LISTENER_GHAYTH_SESSION", "Session — غيث",   "listener (rumman_engine.py)",          ".env",    "TELEGRAM_LISTENER_GHAYTH_SESSION"),
        ("TELEGRAM_BACKFILL_RAWI_SESSION",   "Session — راوي",  "backfill (telegram_backfill_worker)",  ".env",    "TELEGRAM_BACKFILL_RAWI_SESSION"),
        ("TELEGRAM_MEDIA_IBRAHIM_SESSION",   "Session — إبراهيم","media (telegram_download_worker)",    ".env",    "TELEGRAM_MEDIA_IBRAHIM_SESSION"),
        ("TELEGRAM_BOT_TOKEN",               "Bot Token",        "bot (telegram_bot.py)",               "Railway", "TELEGRAM_BOT_TOKEN"),
        ("SUPABASE_URL",                     "Database",         "all workers",                         ".env",    "SUPABASE_URL"),
        ("SUPABASE_KEY",                     "Database",         "all workers",                         ".env",    "SUPABASE_KEY"),
        ("OPENAI_API_KEY",                   "AI / ML",          "embed, search, intelligence, attribution", ".env","OPENAI_API_KEY"),
        ("RAILWAY_API_TOKEN",                "Infrastructure",   "railway CLI (local deploys)",          "Railway", "RAILWAY_API_TOKEN"),
        ("RUMMAN_USER_SALT",                 "Privacy",          "bot, search (user ID hashing)",        "Railway", "RUMMAN_USER_SALT"),
        ("SEU_TENANT_ID",                    "Config",           "all workers",                          "Railway", "SEU_TENANT_ID"),
        ("SEARCH_API_URL",                   "Internal URL",     "bot → search (internal Railway URL)",  "Railway", "SEARCH_API_URL"),
        ("RUMMAN_OPS_CHAT_ID",               "Config",           "weekly_report.py",                     "Railway", "RUMMAN_OPS_CHAT_ID"),
        ("INTELLIGENCE_WORKER_ENABLED",      "Feature Gate",     "intelligence_worker.py",               "Railway", "INTELLIGENCE_WORKER_ENABLED"),
        ("ATTRIBUTION_WORKER_ENABLED",       "Feature Gate",     "attribution_worker.py",                "Railway", "ATTRIBUTION_WORKER_ENABLED"),
        ("LOG_LEVEL",                        "Config",           "all workers",                          ".env",    "LOG_LEVEL"),
    ]

    for var_name, category, used_by, source_label, env_key in creds:
        value = env.get(env_key, _NA)
        is_set = value != _NA
        is_session = "SESSION" in var_name

        data_cell(ws, r, 1, var_name, bold=True, align="left")
        data_cell(ws, r, 2, category, bold=False, align="left")
        data_cell(ws, r, 3, used_by, bold=False, align="left", wrap=True)
        # Source badge
        src_c = ws.cell(row=r, column=4, value=source_label)
        src_c.font = Font(bold=True, size=9,
                          color=C_WHITE if source_label == ".env" else "000000")
        src_c.fill = _fill(C_MID if source_label == ".env" else C_YELLOW)
        src_c.alignment = _align("center", "center")
        src_c.border = _border(style="hair")
        # Value cell
        c = ws.cell(row=r, column=5, value=value)
        c.font = Font(name="Courier New", size=7 if is_session else 9,
                      color="000000" if is_set else C_RED,
                      bold=is_set and not is_session)
        c.fill = _fill("FEF9E7") if is_set else _fill("FDECEA")
        c.alignment = _align("left", "top", wrap=True)
        c.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 72 if is_session else 18
        r += 1

    r += 2

    # ── Rotation guide
    section_header(ws, r, 1, "  HOW TO ROTATE EACH CREDENTIAL TYPE", span=COLS)
    r += 1
    col_header(ws, r, 1, "Type")
    col_header(ws, r, 2, "Steps")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS)
    r += 1
    rotation_guide = [
        ("Telegram Sessions\n(GHAYTH / RAWI / IBRAHIM)",
         "1. Run locally: python3 auth_session.py\n"
         "2. Enter phone (+966XXXXXXXXX) → receive OTP → enter OTP\n"
         "3. Copy the printed StringSession string (one line, starts with 1BJWap...)\n"
         "4. Go to Railway → project → service → Variables\n"
         "5. Update TELEGRAM_LISTENER_GHAYTH_SESSION (or RAWI / IBRAHIM) with new value\n"
         "6. Redeploy the affected process (listener / backfill / media)\n"
         "Warning: never run two processes with the same session simultaneously → AuthKeyDuplicatedError"),
        ("OpenAI API Key",
         "1. platform.openai.com → API keys → Create new secret key\n"
         "2. Copy key immediately (shown once)\n"
         "3. Update OPENAI_API_KEY in local .env AND Railway Variables\n"
         "4. Revoke the old key in OpenAI dashboard\n"
         "5. Monitor ai_runs table for any failures after rotation"),
        ("Supabase Service-Role Key",
         "1. supabase.com → project → Settings → API → service_role key → Reveal\n"
         "2. Update SUPABASE_KEY in local .env AND Railway Variables\n"
         "3. Supabase does not auto-rotate — manual rotation only\n"
         "Note: the SUPABASE_URL never changes for a given project (it's the project ref)"),
        ("Telegram Bot Token",
         "1. Telegram app → @BotFather → /mybots → select bot → API Token → Revoke token\n"
         "2. BotFather generates a new token immediately\n"
         "3. Update TELEGRAM_BOT_TOKEN in Railway Variables\n"
         "4. Redeploy the bot process\n"
         "Note: revoking immediately invalidates the old token — brief downtime expected"),
    ]
    for rot_type, rot_steps in rotation_guide:
        data_cell(ws, r, 1, rot_type, bold=True, wrap=True, align="left")
        c = ws.cell(row=r, column=2, value=rot_steps)
        c.font = Font(size=9, color="000000")
        c.fill = _fill(C_LGRAY) if r % 2 == 0 else _fill(C_WHITE)
        c.alignment = _align("left", "top", wrap=True)
        c.border = _border(style="hair")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS)
        ws.row_dimensions[r].height = 72
        r += 1

    widths = [40, 20, 28, 10, 70, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate RUMMAN master operational workbook")
    parser.add_argument("--out", default=None,
                        help="Output path (default: data/RUMMAN_OPS_YYYY-MM-DD.xlsx)")
    args = parser.parse_args()

    if args.out:
        out_path = Path(args.out)
    else:
        data_dir = Path(__file__).parent / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        out_path = data_dir / f"RUMMAN_OPS_{date.today().isoformat()}.xlsx"

    repo_root = Path(__file__).parent.parent
    env_values = _load_env(repo_root)
    env_loaded = len(env_values) > 0

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheets = [
        ("COMPASS",    lambda ws: build_compass(ws)),
        ("SYSTEMS",    lambda ws: build_systems(ws)),
        ("PROCESSES",  lambda ws: build_processes(ws)),
        ("CORPUS",     lambda ws: build_corpus(ws)),
        ("STRATEGY",   lambda ws: build_strategy(ws)),
        ("OPEN ITEMS", lambda ws: build_open_items(ws)),
        ("KEYS",       lambda ws: build_credentials(ws, env_values)),
    ]

    for name, builder in sheets:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        builder(ws)
        print(f"  Built: {name}")

    wb.save(out_path)
    print(f"\nWorkbook saved → {out_path}")
    print(f"Sheets: {', '.join(s[0] for s in sheets)}")
    if env_loaded:
        print(f"\n  KEYS sheet: populated from .env ({len(env_values)} variables found)")
    else:
        print(f"\n  KEYS sheet: .env not found — all values show as NOT SET")


if __name__ == "__main__":
    main()
